import openpyxl

wb = openpyxl.load_workbook('diningbar selfish 様 HP制作依頼_ヒアリングシート .xlsx', data_only=True)

with open('hearing_sheet_data.txt', 'w', encoding='utf-8') as f:
    for sheet_name in ['ヒアリングシート（一般）', 'ヒアリングシート（飲食）']:
        if sheet_name in wb.sheetnames:
            f.write(f"\n--- {sheet_name} ---\n")
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                if any(val.strip() for val in row_data):
                    while row_data and not row_data[-1].strip():
                        row_data.pop()
                    f.write(" | ".join(row_data) + "\n")
